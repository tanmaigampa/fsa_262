"""
Mapper Module
Maps cleaned financial data to the canonical model using intelligent field matching
"""

import pandas as pd
from typing import Dict, List, Optional

from ..core import CanonicalFinancialData, StatementType
from .field_matcher import FieldMatcher


class FinancialDataMapper:
    """
    Maps cleaned financial data from various formats into the canonical model
    Works with ANY company's financial statements
    """
    
    def __init__(self, cleaned_data: Dict[str, pd.DataFrame], metadata: Dict):
        self.cleaned_data = cleaned_data
        self.metadata = metadata
        self.canonical_model = CanonicalFinancialData()
        self.field_matcher = FieldMatcher()
        self.mapping_log = []  # Track what was mapped
        
    def log_mapping(self, source_field: str, canonical_field: str, statement_type: str):
        """Log a successful mapping"""
        self.mapping_log.append({
            'source': source_field,
            'canonical': canonical_field,
            'statement': statement_type
        })
    
    def map_profit_loss(self, df: pd.DataFrame) -> None:
        """
        Map P&L statement to canonical model
        """
        if df.empty:
            return
        
        statement_type = "P&L"
        years = [col for col in df.columns if col != 'Particulars']
        
        # Update fiscal years in canonical model
        self.canonical_model.fiscal_years = sorted(list(set(
            self.canonical_model.fiscal_years + years
        )))
        
        # Try to map each row to canonical fields
        for _, row in df.iterrows():
            field_name = row['Particulars']
            canonical_field = self.field_matcher.find_canonical_field(field_name)
            
            if canonical_field:
                # Map values for each year
                for year in years:
                    value = row.get(year)
                    if pd.notna(value):
                        self.canonical_model.set_value(canonical_field, year, float(value))
                
                self.log_mapping(field_name, canonical_field, statement_type)
        
        # Compute derived fields
        self._compute_pl_derived_fields(years)
    
    def map_balance_sheet(self, df: pd.DataFrame) -> None:
        """
        Map Balance Sheet to canonical model
        """
        if df.empty:
            return
        
        statement_type = "Balance Sheet"
        years = [col for col in df.columns if col != 'Particulars']
        
        # Update fiscal years
        self.canonical_model.fiscal_years = sorted(list(set(
            self.canonical_model.fiscal_years + years
        )))
        
        # Try to map each row
        for _, row in df.iterrows():
            field_name = row['Particulars']
            canonical_field = self.field_matcher.find_canonical_field(field_name)
            
            if canonical_field:
                for year in years:
                    value = row.get(year)
                    if pd.notna(value):
                        self.canonical_model.set_value(canonical_field, year, float(value))
                
                self.log_mapping(field_name, canonical_field, statement_type)
        
        # Compute derived fields
        self._compute_bs_derived_fields(years)
    
    def map_cash_flow(self, df: pd.DataFrame) -> None:
        """
        Map Cash Flow statement to canonical model
        """
        if df.empty:
            return
        
        statement_type = "Cash Flow"
        years = [col for col in df.columns if col != 'Particulars']
        
        # Update fiscal years
        self.canonical_model.fiscal_years = sorted(list(set(
            self.canonical_model.fiscal_years + years
        )))
        
        # Try to map each row
        for _, row in df.iterrows():
            field_name = row['Particulars']
            canonical_field = self.field_matcher.find_canonical_field(field_name)
            
            if canonical_field:
                for year in years:
                    value = row.get(year)
                    if pd.notna(value):
                        self.canonical_model.set_value(canonical_field, year, float(value))
                
                self.log_mapping(field_name, canonical_field, statement_type)
    
    def _compute_pl_derived_fields(self, years: List[str]) -> None:
        """
        Compute P&L derived fields where possible
        """
        for year in years:
            # Compute gross profit if not present
            revenue = self.canonical_model.get_value('revenue', year)
            cogs = self.canonical_model.get_value('cost_of_goods_sold', year)
            gross_profit = self.canonical_model.get_value('gross_profit', year)
            
            if gross_profit is None and revenue is not None and cogs is not None:
                self.canonical_model.set_value('gross_profit', year, revenue - cogs)
            
            # Compute EBIT if not present
            ebit = self.canonical_model.get_value('ebit', year)
            ebitda = self.canonical_model.get_value('ebitda', year)
            dep_amort = self.canonical_model.get_value('depreciation_amortization', year)
            
            if ebit is None and ebitda is not None and dep_amort is not None:
                self.canonical_model.set_value('ebit', year, ebitda - dep_amort)
    
    def _compute_bs_derived_fields(self, years: List[str]) -> None:
        """
        Compute Balance Sheet derived fields where possible
        """
        for year in years:
            # Compute total equity if not present
            total_equity = self.canonical_model.get_value('total_equity', year)
            share_capital = self.canonical_model.get_value('share_capital', year)
            reserves = self.canonical_model.get_value('reserves_surplus', year)
            
            if total_equity is None and share_capital is not None and reserves is not None:
                computed_equity = share_capital + reserves
                self.canonical_model.set_value('total_equity', year, computed_equity)
            
            # Compute total assets if not present
            total_assets = self.canonical_model.get_value('total_assets', year)
            current_assets = self.canonical_model.get_value('current_assets', year)
            non_current_assets = self.canonical_model.get_value('non_current_assets', year)
            
            if total_assets is None and current_assets is not None and non_current_assets is not None:
                self.canonical_model.set_value('total_assets', year, current_assets + non_current_assets)
            
            # Compute total liabilities if not present
            total_liab = self.canonical_model.get_value('total_liabilities', year)
            current_liab = self.canonical_model.get_value('current_liabilities', year)
            non_current_liab = self.canonical_model.get_value('non_current_liabilities', year)
            
            if total_liab is None and current_liab is not None and non_current_liab is not None:
                self.canonical_model.set_value('total_liabilities', year, current_liab + non_current_liab)
            
            # Compute total debt if not present
            total_debt = self.canonical_model.get_value('total_debt', year)
            short_debt = self.canonical_model.get_value('short_term_debt', year)
            long_debt = self.canonical_model.get_value('long_term_debt', year)
            
            if total_debt is None and short_debt is not None and long_debt is not None:
                self.canonical_model.set_value('total_debt', year, short_debt + long_debt)
            
            # Compute working capital
            working_capital = self.canonical_model.get_value('working_capital', year)
            current_assets = self.canonical_model.get_value('current_assets', year)
            current_liab = self.canonical_model.get_value('current_liabilities', year)
            
            if working_capital is None and current_assets is not None and current_liab is not None:
                self.canonical_model.set_value('working_capital', year, current_assets - current_liab)
    
    def map_all_statements(self) -> CanonicalFinancialData:
        """
        Map all available statements to canonical model
        Returns the populated canonical model
        """
        # Set metadata
        self.canonical_model.company_name = self.metadata.get('company_name', 'Unknown')
        
        # Map each statement type
        if 'P&L' in self.cleaned_data:
            self.map_profit_loss(self.cleaned_data['P&L'])
        
        if 'Balance Sheet' in self.cleaned_data:
            self.map_balance_sheet(self.cleaned_data['Balance Sheet'])
        
        if 'Cash Flow' in self.cleaned_data:
            self.map_cash_flow(self.cleaned_data['Cash Flow'])
        
        return self.canonical_model
    
    def get_mapping_summary(self) -> Dict:
        """
        Get summary of what was successfully mapped
        """
        summary = {
            'total_mappings': len(self.mapping_log),
            'by_statement': {},
            'unmapped_fields': []
        }
        
        # Count by statement type
        for log_entry in self.mapping_log:
            stmt = log_entry['statement']
            if stmt not in summary['by_statement']:
                summary['by_statement'][stmt] = []
            summary['by_statement'][stmt].append(
                f"{log_entry['source']} → {log_entry['canonical']}"
            )
        
        # Find unmapped fields
        for stmt_type, df in self.cleaned_data.items():
            for field in df['Particulars'].unique():
                if not any(log['source'] == field for log in self.mapping_log):
                    summary['unmapped_fields'].append(f"{stmt_type}: {field}")
        
        return summary
    
    def export_canonical_to_excel(self, output_path: str):
        """
        Export canonical model to Excel for verification
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        years = self.canonical_model.get_available_years()
        
        # P&L Sheet
        pl_sheet = wb.create_sheet("P&L")
        pl_fields = [
            'revenue', 'cost_of_goods_sold', 'gross_profit', 'operating_expenses',
            'ebitda', 'depreciation_amortization', 'ebit', 'interest_expense',
            'profit_before_tax', 'tax_expense', 'net_profit'
        ]
        self._create_statement_sheet(pl_sheet, pl_fields, years, "Income Statement")
        
        # Balance Sheet
        bs_sheet = wb.create_sheet("Balance Sheet")
        bs_fields = [
            'current_assets', 'non_current_assets', 'total_assets',
            'current_liabilities', 'non_current_liabilities', 'total_liabilities',
            'share_capital', 'reserves_surplus', 'total_equity',
            'short_term_debt', 'long_term_debt', 'total_debt', 'working_capital'
        ]
        self._create_statement_sheet(bs_sheet, bs_fields, years, "Balance Sheet")
        
        # Cash Flow
        cf_sheet = wb.create_sheet("Cash Flow")
        cf_fields = [
            'operating_cash_flow', 'investing_cash_flow', 'financing_cash_flow',
            'net_cash_flow', 'cash_beginning', 'cash_ending', 'capex'
        ]
        self._create_statement_sheet(cf_sheet, cf_fields, years, "Cash Flow Statement")
        
        wb.save(output_path)
        print(f"Canonical model exported to {output_path}")
    
    def _create_statement_sheet(self, sheet, fields, years, title):
        """Helper to create a formatted statement sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Title
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:E1')
        
        # Company name
        sheet['A2'] = self.canonical_model.company_name
        sheet['A2'].font = Font(italic=True)
        
        # Headers
        sheet['A4'] = 'Particulars'
        sheet['A4'].font = Font(bold=True)
        
        for idx, year in enumerate(years):
            col = chr(66 + idx)  # B, C, D, etc.
            sheet[f'{col}4'] = year
            sheet[f'{col}4'].font = Font(bold=True)
            sheet[f'{col}4'].alignment = Alignment(horizontal='center')
        
        # Data rows
        row_num = 5
        for field in fields:
            # Field name (convert underscore to title case)
            field_display = field.replace('_', ' ').title()
            sheet[f'A{row_num}'] = field_display
            
            # Values for each year
            for idx, year in enumerate(years):
                col = chr(66 + idx)
                value = self.canonical_model.get_value(field, year)
                if value is not None:
                    sheet[f'{col}{row_num}'] = value
                    sheet[f'{col}{row_num}'].number_format = '#,##0.00'
                else:
                    sheet[f'{col}{row_num}'] = '-'
            
            row_num += 1
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 30
        for idx in range(len(years)):
            col = chr(66 + idx)
            sheet.column_dimensions[col].width = 15
